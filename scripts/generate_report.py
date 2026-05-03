#!/usr/bin/env python3
"""
RL-MEC Benchmark Results Report Generator
Generates comprehensive experiment analysis report and data export
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List

try:
    import pandas as pd
except ModuleNotFoundError:
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ModuleNotFoundError:
    Workbook = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    dataframe_to_rows = None

PROJECT_ROOT = Path(__file__).parent.parent

# File paths
RESULTS_DIR = PROJECT_ROOT / "results"
LOGS_DIR = PROJECT_ROOT / "logs"
SCRIPTS_DIR = Path(__file__).parent
FIGURES_DIR = PROJECT_ROOT / "figures"

# Input files
JSON_FILE = RESULTS_DIR / "benchmark_full_500k_20260423_101541.json"
LOG_FILE = LOGS_DIR / "benchmark_full_500k_20260423_101541.log"

# Output files
REPORT_FILE = RESULTS_DIR / "benchmark_report_20260423.md"
ANALYSIS_FILE = RESULTS_DIR / "benchmark_analysis_20260423.xlsx"


class BenchmarkAnalyzer:
    """Analyze benchmark results and generate reports"""

    def __init__(self, json_path: Path):
        """Initialize analyzer with JSON results"""
        if pd is None:
            raise RuntimeError("pandas is required to generate benchmark reports")
        with open(json_path, 'r') as f:
            self.raw_data = json.load(f)

        self.df = self._prepare_dataframe()
        self.benchmark_time = "2026-04-23 10:15:41"
        self.total_timesteps = 500000

    def _prepare_dataframe(self) -> pd.DataFrame:
        """Convert JSON data to DataFrame with categorization"""
        rows = []

        def first_present(entry: Dict, *keys):
            for key in keys:
                value = entry.get(key)
                if value is not None:
                    return value
            return None

        for entry in self.raw_data:
            algo = entry.get('algorithm')
            status = entry.get('status', 'unknown')

            if status == 'ok':
                e2e_latency = first_present(
                    entry,
                    'final_e2e_latency_mean_mean',
                    'final_latency_per_task_mean_mean',
                    'final_latency_mean_mean',
                )
                latency_total = first_present(
                    entry,
                    'final_latency_total_mean_mean',
                    'final_latency_mean_mean',
                )
                rows.append({
                    'algorithm': algo,
                    'environment': entry.get('environment', 'N/A'),
                    'status': status,
                    'reward_mean': entry.get('final_reward_mean_mean'),
                    'reward_std': entry.get('final_reward_std_mean'),
                    'latency_mean': e2e_latency,
                    'e2e_latency_mean': e2e_latency,
                    'e2e_latency_p95': entry.get('final_e2e_latency_p95_mean'),
                    'deadline_miss_rate': entry.get('final_deadline_miss_rate_mean'),
                    'throughput_tasks_per_step': entry.get('final_throughput_tasks_per_step_mean'),
                    'comm_score': entry.get('final_comm_score_mean'),
                    'latency_total': latency_total,
                    'energy_mean': entry.get('final_energy_mean_mean'),
                    'energy_per_task': entry.get('final_energy_per_task_mean_mean'),
                    'train_time_sec': entry.get('train_time_seconds_mean'),
                    'train_timesteps': entry.get('train_timesteps_mean'),
                    'latency_per_step': entry.get('final_latency_per_step_mean_mean'),
                    'latency_per_task': entry.get('final_latency_per_task_mean_mean'),
                    'energy_per_step': entry.get('final_energy_per_step_mean_mean'),
                    'episodes': entry.get('total_episodes_mean'),
                    'updates': entry.get('total_updates_mean'),
                })
            else:
                # Failed run
                error = entry.get('errors', [{}])[0].get('error', 'Unknown error')
                rows.append({
                    'algorithm': algo,
                    'environment': entry.get('environment', 'N/A'),
                    'status': status,
                    'reward_mean': None,
                    'reward_std': None,
                    'latency_mean': None,
                    'e2e_latency_mean': None,
                    'e2e_latency_p95': None,
                    'deadline_miss_rate': None,
                    'throughput_tasks_per_step': None,
                    'comm_score': None,
                    'latency_total': None,
                    'energy_mean': None,
                    'energy_per_task': None,
                    'train_time_sec': None,
                    'train_timesteps': None,
                    'latency_per_step': None,
                    'latency_per_task': None,
                    'energy_per_step': None,
                    'episodes': None,
                    'updates': None,
                    'error': error,
                })

        df = pd.DataFrame(rows)

        # Add algorithm category
        on_policy = ['GRPO', 'PPO', 'A3C', 'TRPO', 'SimPO', 'COMA', 'IPPO', 'MAPPO']
        off_policy = ['SAC', 'DDPG', 'TD3', 'DDQN', 'VDN', 'IQL', 'MADDPG', 'MATD3', 'QMIX']

        df['category'] = df['algorithm'].apply(
            lambda x: 'On-Policy' if x in on_policy else
                      'Off-Policy' if x in off_policy else 'Unknown'
        )

        # Add multi-agent flag
        multi_agent = ['MAPPO', 'QMIX', 'COMA', 'IPPO', 'VDN', 'MADDPG', 'IQL', 'MATD3']
        df['is_multi_agent'] = df['algorithm'].isin(multi_agent)

        # Calculate efficiency metrics for successful runs
        df['reward_per_sec'] = df.apply(
            lambda x: x['reward_mean'] / x['train_time_sec'] if x['status'] == 'ok' and x['train_time_sec'] > 0 else None,
            axis=1
        )

        return df

    def get_success_summary(self) -> Dict:
        """Get success/failure summary"""
        total = len(self.df)
        successful = len(self.df[self.df['status'] == 'ok'])
        failed = total - successful

        return {
            'total': total,
            'successful': successful,
            'failed': failed,
            'success_rate': f"{100 * successful / total:.1f}%"
        }

    def get_ranking(self, metric: str, ascending: bool = False, limit: int = 3) -> List[Dict]:
        """Get top N algorithms by metric"""
        df_ok = self.df[self.df['status'] == 'ok'].copy()
        if df_ok.empty:
            return []

        ranked = df_ok.dropna(subset=[metric]).sort_values(metric, ascending=ascending).head(limit)
        result = []

        for rank, (_, row) in enumerate(ranked.iterrows(), 1):
            result.append({
                'rank': rank,
                'algorithm': row['algorithm'],
                'environment': row['environment'].split('-')[-2:],  # Get env suffix
                'value': round(row[metric], 4),
                'category': row['category'],
            })

        return result

    def get_category_stats(self, category: str) -> Dict:
        """Get statistics for algorithm category"""
        df_cat = self.df[(self.df['category'] == category) & (self.df['status'] == 'ok')]

        if df_cat.empty:
            return {'count': 0}

        return {
            'count': len(df_cat),
            'algorithms': df_cat['algorithm'].tolist(),
            'avg_reward': df_cat['reward_mean'].mean(),
            'avg_time': df_cat['train_time_sec'].mean(),
            'avg_latency': df_cat['latency_mean'].mean(),
            'avg_energy': df_cat['energy_mean'].mean(),
        }

    def get_environment_stats(self) -> Dict:
        """Get statistics by environment type"""
        stats = {}
        for env in self.df['environment'].unique():
            df_env = self.df[(self.df['environment'] == env) & (self.df['status'] == 'ok')]
            stats[env] = {
                'count': len(df_env),
                'algorithms': df_env['algorithm'].tolist(),
                'avg_reward': df_env['reward_mean'].mean(),
                'best_reward': df_env['reward_mean'].max(),
                'best_algo': df_env.loc[df_env['reward_mean'].idxmax(), 'algorithm'],
            }
        return stats


def _append_composite_ranking_section(analyzer: BenchmarkAnalyzer, lines: List[str]) -> None:
    """Append the 综合评分排名 section to the report lines.

    Reads ``composite_scores`` and ``robustness`` fields from the raw JSON
    data.  Gracefully skips algorithms that lack these fields.
    """
    profiles = ["balanced", "latency_critical", "energy_constrained"]
    profile_labels = {
        "balanced": "Balanced (均衡)",
        "latency_critical": "Latency-Critical (时延优先)",
        "energy_constrained": "Energy-Constrained (能耗优先)",
    }

    # Collect algorithms that have composite_scores
    algo_data: Dict[str, Dict] = {}
    for entry in analyzer.raw_data:
        algo = entry.get("algorithm", "unknown")
        cs = entry.get("composite_scores")
        if cs and isinstance(cs, dict):
            algo_data[algo] = {
                "composite_scores": cs,
                "robustness": entry.get("robustness"),
            }

    if not algo_data:
        return  # No composite data available — skip section entirely

    lines.append("## 综合评分排名")
    lines.append("")
    lines.append("基于多维度加权归一化的综合评分，从奖励、时延、能耗、稳定性四个维度对算法进行综合排名。")
    lines.append("三个评分配置（Profile）分别侧重不同优化目标：均衡、时延优先、能耗优先。")
    lines.append("")

    # ── Per-profile ranking tables ─────────────────────────────────────
    for profile in profiles:
        label = profile_labels.get(profile, profile)
        lines.append(f"### {label}")
        lines.append("")

        # Collect (algo, score_data) for this profile
        profile_entries = []
        for algo, data in algo_data.items():
            ps = data["composite_scores"].get(profile)
            if ps and isinstance(ps, dict) and "score" in ps:
                profile_entries.append((algo, ps))

        if not profile_entries:
            lines.append(f"*暂无 {profile} profile 数据*")
            lines.append("")
            continue

        # Sort by composite score descending
        profile_entries.sort(key=lambda x: x[1]["score"], reverse=True)

        lines.append("| Rank | Algorithm | Composite Score | Reward | Latency | Energy | Stability |")
        lines.append("|------|-----------|----------------|--------|---------|--------|-----------|")

        for rank, (algo, ps) in enumerate(profile_entries, 1):
            score = ps["score"]
            bd = ps.get("breakdown", {})
            reward_bd = bd.get("reward", 0.0)
            latency_bd = bd.get("latency", 0.0)
            energy_bd = bd.get("energy", 0.0)
            stability_bd = bd.get("stability", 0.0)
            lines.append(
                f"| {rank} | {algo} | {score:.4f} | {reward_bd:.4f} | "
                f"{latency_bd:.4f} | {energy_bd:.4f} | {stability_bd:.4f} |"
            )

        lines.append("")

    # ── Robustness summary table ───────────────────────────────────────
    lines.append("### 鲁棒性分析 (Robustness)")
    lines.append("")
    lines.append("综合各 Profile 排名，评估算法在不同优化目标下的稳定表现。")
    lines.append("\"Robust\" 表示该算法在所有 Profile 中均排名前 3。")
    lines.append("")

    robustness_entries = []
    for algo, data in algo_data.items():
        rob = data.get("robustness")
        if rob and isinstance(rob, dict):
            robustness_entries.append((algo, rob))

    if robustness_entries:
        # Sort by avg_rank ascending
        robustness_entries.sort(key=lambda x: x[1].get("avg_rank", 999))

        lines.append("| Algorithm | Avg Rank | Worst Rank | Best Rank | Robust |")
        lines.append("|-----------|----------|------------|-----------|--------|")

        for algo, rob in robustness_entries:
            avg_rank = rob.get("avg_rank", 0)
            worst_rank = rob.get("worst_rank", 0)
            best_rank = rob.get("best_rank", 0)
            is_robust = rob.get("robust", False)
            robust_str = "✅" if is_robust else "❌"
            lines.append(
                f"| {algo} | {avg_rank:.2f} | {worst_rank} | {best_rank} | {robust_str} |"
            )

        lines.append("")

    # ── Figure references ──────────────────────────────────────────────
    lines.append("### 相关图表")
    lines.append("")
    lines.append("- 综合排名柱状图: `figures/composite_ranking.png` / `figures/composite_ranking.pdf`")
    lines.append("- Top-6 雷达图: `figures/radar_top6.png` / `figures/radar_top6.pdf`")
    lines.append("")


def _display_path(path: Path) -> str:
    """Return a project-relative display path when possible."""
    try:
        return path.resolve().relative_to(PROJECT_ROOT.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def _find_convergence_figure(figures_dir: Path) -> Path | None:
    """Prefer clean convergence figures, then raw/legacy fallbacks."""
    candidates = [
        figures_dir / "convergence_curves_clean_all.png",
        figures_dir / "convergence_curves_clean_all.pdf",
        figures_dir / "convergence_curves_raw_all.png",
        figures_dir / "convergence_curves_raw_all.pdf",
        figures_dir / "convergence_curves.png",
        figures_dir / "convergence_curves.pdf",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _load_convergence_quality_records(figures_dir: Path) -> list[dict]:
    """Load convergence quality JSON records if present."""
    report_path = figures_dir / "convergence_quality_report.json"
    if not report_path.exists():
        return []
    try:
        with open(report_path, encoding="utf-8") as f:
            records = json.load(f)
    except (OSError, json.JSONDecodeError):
        return []
    return records if isinstance(records, list) else []


def _append_convergence_quality_section(
    lines: List[str],
    *,
    figures_dir: Path,
    convergence_figure: Path | None,
    quality_records: list[dict],
) -> None:
    """Append convergence figure and quality-report notes."""
    quality_md = figures_dir / "convergence_quality_report.md"
    if convergence_figure is None and not quality_md.exists() and not quality_records:
        return

    lines.append("## Convergence Quality Notes")
    lines.append("")
    if convergence_figure is not None:
        lines.append(f"- Preferred convergence figure: `{_display_path(convergence_figure)}`")
        if convergence_figure.name.startswith("convergence_curves_clean_all"):
            lines.append("- This report uses the clean publication figure when available.")
    raw_png = figures_dir / "convergence_curves_raw_all.png"
    raw_pdf = figures_dir / "convergence_curves_raw_all.pdf"
    if raw_png.exists() or raw_pdf.exists():
        raw_path = raw_png if raw_png.exists() else raw_pdf
        lines.append(f"- Raw diagnostic figure: `{_display_path(raw_path)}`")
    if quality_md.exists():
        lines.append(f"- Full quality report: `{_display_path(quality_md)}`")
    lines.append("")

    warnings = []
    for record in quality_records:
        ratio = record.get("outlier_ratio") or 0.0
        if (
            record.get("severe_outlier")
            or ratio >= 0.20
            or record.get("skipped_from_clean_plot")
            or record.get("excluded_from_clean_plot")
        ):
            warnings.append(record)

    if not warnings:
        lines.append("No severe convergence quality warnings were reported.")
        lines.append("")
        return

    lines.append("| Algorithm | Seed | Metric | Warning | Outlier Ratio |")
    lines.append("|-----------|------|--------|---------|---------------|")
    for record in warnings[:20]:
        warning_bits = []
        if record.get("severe_outlier"):
            warning_bits.append("severe_outlier")
        if record.get("skipped_from_clean_plot"):
            warning_bits.append(record.get("skip_reason", "skipped_from_clean_plot"))
        if record.get("excluded_from_clean_plot"):
            warning_bits.append(record.get("reason", "excluded_from_clean_plot"))
        warning = "; ".join(str(bit) for bit in warning_bits) or "outlier_ratio"
        lines.append(
            "| {algorithm} | {seed} | {metric} | {warning} | {ratio:.3f} |".format(
                algorithm=record.get("algorithm", "unknown"),
                seed=record.get("seed", "unknown"),
                metric=record.get("metric", "unknown"),
                warning=warning,
                ratio=float(record.get("outlier_ratio") or 0.0),
            )
        )
    lines.append("")


def generate_markdown_report(
    analyzer: BenchmarkAnalyzer,
    output_file: Path,
    figures_dir: Path | None = None,
):
    """Generate comprehensive markdown report"""

    lines = []
    figures_dir = figures_dir or FIGURES_DIR
    convergence_figure = _find_convergence_figure(figures_dir)
    quality_records = _load_convergence_quality_records(figures_dir)

    # Header
    lines.append("# RL-MEC Benchmark Results Report")
    lines.append("**Experiment Date**: 2026-04-23 10:15:41")
    lines.append(f"**Report Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("**Total Timesteps**: 500,000 (500k)")
    lines.append("")

    # Table of Contents
    lines.append("## Table of Contents")
    lines.append("- [Executive Summary](#executive-summary)")
    lines.append("- [Performance Rankings](#performance-rankings)")
    lines.append("- [Algorithm Category Analysis](#algorithm-category-analysis)")
    lines.append("- [Environment Analysis](#environment-analysis)")
    lines.append("- [Detailed Results](#detailed-results)")
    lines.append("- [Key Findings](#key-findings)")
    lines.append("- [Failure Diagnosis](#failure-diagnosis)")
    lines.append("- [Recommendations](#recommendations)")
    if convergence_figure is not None or quality_records or (figures_dir / "convergence_quality_report.md").exists():
        lines.append("- [Convergence Quality Notes](#convergence-quality-notes)")
    lines.append("- [综合评分排名](#综合评分排名)")
    lines.append("")

    # Executive Summary
    lines.append("## Executive Summary")
    lines.append("")

    summary = analyzer.get_success_summary()
    lines.append(f"This benchmark evaluates **{summary['total']} RL algorithms** on GameTheory-based MEC environments:")
    lines.append(f"- **Success Rate**: {summary['successful']}/{summary['total']} ({summary['success_rate']})")
    lines.append(f"- **Failed**: {summary['failed']} algorithm(s) (IPPO - configuration error)")
    lines.append("")

    # Top performers
    lines.append("### Top Performers")
    lines.append("")

    top_reward = analyzer.get_ranking('reward_mean', ascending=False, limit=3)
    top_time = analyzer.get_ranking('train_time_sec', ascending=True, limit=3)
    latency_rank_metric = 'e2e_latency_p95'
    if analyzer.df[latency_rank_metric].dropna().empty:
        latency_rank_metric = 'e2e_latency_mean'
    top_latency = analyzer.get_ranking(latency_rank_metric, ascending=True, limit=3)
    top_comm = analyzer.get_ranking('comm_score', ascending=False, limit=3)

    if top_reward:
        lines.append(f"**Highest Reward**: {top_reward[0]['algorithm']} ({top_reward[0]['value']:.4f})")
    if top_time:
        lines.append(f"**Fastest Training**: {top_time[0]['algorithm']} ({top_time[0]['value']:.1f}s)")
    if top_latency:
        lines.append(f"**Lowest E2E Latency**: {top_latency[0]['algorithm']} ({top_latency[0]['value']:.4f})")
    if top_comm:
        lines.append(f"**Best Communication Score**: {top_comm[0]['algorithm']} ({top_comm[0]['value']:.4f})")

    lines.append("")

    # Performance Rankings
    lines.append("## Performance Rankings")
    lines.append("")

    lines.append("### Reward Rankings (Top 3)")
    lines.append("| Rank | Algorithm | Environment | Reward | Category |")
    lines.append("|------|-----------|-------------|--------|----------|")
    for r in top_reward:
        lines.append(f"| {r['rank']} | {r['algorithm']} | {'-'.join(r['environment'])} | {r['value']:.4f} | {r['category']} |")
    lines.append("")

    lines.append("### Training Time Rankings (Top 3 - Fastest)")
    lines.append("| Rank | Algorithm | Time (sec) | Environment |")
    lines.append("|------|-----------|-----------|-------------|")
    for r in top_time:
        lines.append(f"| {r['rank']} | {r['algorithm']} | {r['value']:.1f} | {'-'.join(r['environment'])} |")
    lines.append("")

    lines.append("### E2E Latency Rankings (Top 3 - Lowest)")
    lines.append("| Rank | Algorithm | E2E Latency | Environment |")
    lines.append("|------|-----------|---------|-------------|")
    for r in top_latency:
        lines.append(f"| {r['rank']} | {r['algorithm']} | {r['value']:.4f} | {'-'.join(r['environment'])} |")
    lines.append("")

    if top_comm:
        lines.append("### Communication Score Rankings (Top 3)")
        lines.append("| Rank | Algorithm | Comm Score | Environment |")
        lines.append("|------|-----------|------------|-------------|")
        for r in top_comm:
            lines.append(f"| {r['rank']} | {r['algorithm']} | {r['value']:.4f} | {'-'.join(r['environment'])} |")
        lines.append("")

    # Algorithm Category Analysis
    lines.append("## Algorithm Category Analysis")
    lines.append("")

    for category in ['On-Policy', 'Off-Policy']:
        stats = analyzer.get_category_stats(category)
        if stats['count'] > 0:
            lines.append(f"### {category} Algorithms ({stats['count']} algorithms)")
            lines.append(f"- **Algorithms**: {', '.join(stats['algorithms'])}")
            lines.append(f"- **Avg Reward**: {stats['avg_reward']:.4f}")
            lines.append(f"- **Avg Training Time**: {stats['avg_time']:.1f}s")
            lines.append(f"- **Avg Latency**: {stats['avg_latency']:.4f}")
            lines.append(f"- **Avg Energy**: {stats['avg_energy']:.4f}")
            lines.append("")

    # Environment Analysis
    lines.append("## Environment Analysis")
    lines.append("")

    env_stats = analyzer.get_environment_stats()
    for env, stats in env_stats.items():
        env_name = env.split('MEC-v1-')[1] if 'MEC-v1-' in env else env
        lines.append(f"### {env_name}")
        lines.append(f"- **Algorithm Count**: {stats['count']}")
        lines.append(f"- **Best Algorithm**: {stats['best_algo']} (reward: {stats['best_reward']:.4f})")
        lines.append(f"- **Average Reward**: {stats['avg_reward']:.4f}")
        lines.append("")

    # Detailed Results Table
    lines.append("## Detailed Results")
    lines.append("")

    df_ok = analyzer.df[analyzer.df['status'] == 'ok'].sort_values('e2e_latency_mean', ascending=True)
    lines.append("| Algorithm | Environment | E2E Latency | P95 | Deadline Miss | Throughput | Comm Score | Total Latency | Reward | Energy | Time (s) |")
    lines.append("|-----------|-------------|-------------|-----|---------------|------------|------------|---------------|--------|--------|----------|")

    for _, row in df_ok.iterrows():
        env_short = row['environment'].split('-')[-1]
        p95 = row['e2e_latency_p95'] if pd.notna(row['e2e_latency_p95']) else row['e2e_latency_mean']
        miss = row['deadline_miss_rate'] if pd.notna(row['deadline_miss_rate']) else 0.0
        throughput = row['throughput_tasks_per_step'] if pd.notna(row['throughput_tasks_per_step']) else 0.0
        comm_score = row['comm_score'] if pd.notna(row['comm_score']) else 0.0
        latency_total = row['latency_total'] if pd.notna(row['latency_total']) else row['latency_mean']
        lines.append(
            f"| {row['algorithm']} | {env_short} | {row['e2e_latency_mean']:.4f} | {p95:.4f} | "
            f"{miss:.4f} | {throughput:.4f} | {comm_score:.4f} | {latency_total:.4f} | "
            f"{row['reward_mean']:.4f} | {row['energy_mean']:.4f} | {row['train_time_sec']:.1f} |"
        )
    lines.append("")

    # Key Findings
    lines.append("## Key Findings")
    lines.append("")

    lines.append("### Convergence Speed")
    lines.append(f"- **Fastest convergence**: {analyzer.get_ranking('train_time_sec', ascending=True, limit=1)[0]['algorithm']} ({analyzer.get_ranking('train_time_sec', ascending=True, limit=1)[0]['value']:.1f}s)")
    lines.append("- Multi-agent algorithms (MADDPG, MATD3, QMIX) require significantly longer training time")
    lines.append("")

    lines.append("### Sample Efficiency")
    reward_per_sec = analyzer.df[analyzer.df['status'] == 'ok'].nlargest(3, 'reward_per_sec')
    if not reward_per_sec.empty:
        lines.append("Top performers by reward per second:")
        for algo, val in reward_per_sec[['algorithm', 'reward_per_sec']].values:
            lines.append(f"- {algo}: {val:.4f} reward/sec")
    lines.append("")

    lines.append("### Algorithm Suitability")
    lines.append("- **Continuous Action Space**: GRPO, PPO, and MADDPG show strong performance")
    lines.append("- **Discrete Action Space**: QMIX and MAPPO show highest rewards among discrete algorithms")
    lines.append("- **Single-Agent**: On-policy algorithms (GRPO, PPO) converge faster")
    lines.append("- **Multi-Agent**: MADDPG and MATD3 achieve highest absolute rewards")
    lines.append("")

    # Failure Diagnosis
    lines.append("## Failure Diagnosis")
    lines.append("")

    failed = analyzer.df[analyzer.df['status'] == 'failed']
    if not failed.empty:
        lines.append("### Failed Algorithms")
        lines.append("")
        for _, row in failed.iterrows():
            lines.append(f"**{row['algorithm']} ({row['environment']})**")
            lines.append(f"- Error: `{row.get('error', 'Unknown error')}`")
            lines.append("- Recommendation: Check configuration parameter validation in IPPOAgent initialization")
            lines.append("")

    # Recommendations
    lines.append("## Recommendations")
    lines.append("")
    lines.append("### For Production Deployment")
    lines.append("1. **Use GRPO or PPO** for continuous control (fast, reliable, good reward)")
    lines.append("2. **Use MADDPG/MATD3** for multi-agent continuous tasks (highest rewards)")
    lines.append("3. **Use MAPPO/QMIX** for discrete multi-agent tasks")
    lines.append("")

    lines.append("### For Future Experiments")
    lines.append("1. **Fix IPPO**: Remove or properly handle 'use_game_theory' parameter")
    lines.append("2. **Hyperparameter Tuning**: Current configs are defaults; tuning could improve SAC and DDQN")
    lines.append("3. **Extended Training**: Run to 1M+ timesteps to see convergence patterns")
    lines.append("4. **Multi-seed Analysis**: Current results are single-seed (seed=42); use 3+ seeds for statistical significance")
    lines.append("")

    _append_convergence_quality_section(
        lines,
        figures_dir=figures_dir,
        convergence_figure=convergence_figure,
        quality_records=quality_records,
    )

    # ── Composite Score Ranking (综合评分排名) ──────────────────────────
    _append_composite_ranking_section(analyzer, lines)

    # Write file
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"OK Markdown report generated: {output_file}")


def _add_composite_scores_sheet(analyzer: BenchmarkAnalyzer, wb: Workbook) -> None:
    """Add a 'Composite Scores' sheet to the workbook.

    Creates two sub-tables:
    1. Per-profile ranking (balanced / latency_critical / energy_constrained)
    2. Robustness summary
    """
    profiles = ["balanced", "latency_critical", "energy_constrained"]

    # Collect algorithms that have composite_scores
    algo_data: Dict[str, Dict] = {}
    for entry in analyzer.raw_data:
        algo = entry.get("algorithm", "unknown")
        cs = entry.get("composite_scores")
        if cs and isinstance(cs, dict):
            algo_data[algo] = {
                "composite_scores": cs,
                "robustness": entry.get("robustness"),
            }

    if not algo_data:
        return  # No composite data — skip sheet

    ws = wb.create_sheet("Composite Scores")

    # ── Header styles ──────────────────────────────────────────────────
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    current_row = 1

    # ── Per-profile ranking tables ─────────────────────────────────────
    profile_labels = {
        "balanced": "Balanced",
        "latency_critical": "Latency-Critical",
        "energy_constrained": "Energy-Constrained",
    }

    for profile in profiles:
        label = profile_labels.get(profile, profile)

        # Section title
        ws.cell(row=current_row, column=1, value=f"Profile: {label}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        # Header row
        headers = ["Rank", "Algorithm", "Composite Score", "Reward", "Latency", "Energy", "Stability"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Collect and sort entries
        profile_entries = []
        for algo, data in algo_data.items():
            ps = data["composite_scores"].get(profile)
            if ps and isinstance(ps, dict) and "score" in ps:
                profile_entries.append((algo, ps))

        profile_entries.sort(key=lambda x: x[1]["score"], reverse=True)

        for rank, (algo, ps) in enumerate(profile_entries, 1):
            bd = ps.get("breakdown", {})
            ws.cell(row=current_row, column=1, value=rank)
            ws.cell(row=current_row, column=2, value=algo)
            ws.cell(row=current_row, column=3, value=round(ps["score"], 6))
            ws.cell(row=current_row, column=4, value=round(bd.get("reward", 0.0), 6))
            ws.cell(row=current_row, column=5, value=round(bd.get("latency", 0.0), 6))
            ws.cell(row=current_row, column=6, value=round(bd.get("energy", 0.0), 6))
            ws.cell(row=current_row, column=7, value=round(bd.get("stability", 0.0), 6))
            # Number format for score columns
            for c in range(3, 8):
                ws.cell(row=current_row, column=c).number_format = "0.000000"
            current_row += 1

        current_row += 1  # Blank row between profiles

    # ── Robustness summary table ───────────────────────────────────────
    ws.cell(row=current_row, column=1, value="Robustness Summary")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1

    rob_headers = ["Algorithm", "Avg Rank", "Worst Rank", "Best Rank", "Rank Variance", "Robust"]
    for c_idx, h in enumerate(rob_headers, 1):
        cell = ws.cell(row=current_row, column=c_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    current_row += 1

    robustness_entries = []
    for algo, data in algo_data.items():
        rob = data.get("robustness")
        if rob and isinstance(rob, dict):
            robustness_entries.append((algo, rob))

    robustness_entries.sort(key=lambda x: x[1].get("avg_rank", 999))

    for algo, rob in robustness_entries:
        ws.cell(row=current_row, column=1, value=algo)
        ws.cell(row=current_row, column=2, value=round(rob.get("avg_rank", 0), 4))
        ws.cell(row=current_row, column=3, value=rob.get("worst_rank", 0))
        ws.cell(row=current_row, column=4, value=rob.get("best_rank", 0))
        ws.cell(row=current_row, column=5, value=round(rob.get("rank_variance", 0), 4))
        ws.cell(row=current_row, column=6, value="Yes" if rob.get("robust", False) else "No")
        ws.cell(row=current_row, column=2).number_format = "0.0000"
        ws.cell(row=current_row, column=5).number_format = "0.0000"
        current_row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_excel_analysis(analyzer: BenchmarkAnalyzer, output_file: Path):
    """Generate Excel workbook with detailed analysis"""

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Raw Results
    ws1 = wb.create_sheet("Raw Results")
    df_all = analyzer.df.copy()
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)

    # Auto-adjust columns
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: Performance Rankings
    ws2 = wb.create_sheet("Rankings")

    df_ok = analyzer.df[analyzer.df['status'] == 'ok']
    rankings_data = []

    # Reward ranking
    for rank, (_, row) in enumerate(df_ok.nlargest(10, 'reward_mean').iterrows(), 1):
        rankings_data.append(['Reward', rank, row['algorithm'], row['reward_mean']])

    # Time ranking
    for rank, (_, row) in enumerate(df_ok.nsmallest(10, 'train_time_sec').iterrows(), 1):
        rankings_data.append(['Time (Fast)', rank, row['algorithm'], row['train_time_sec']])

    # E2E latency ranking. Prefer P95 when available, otherwise mean E2E latency.
    latency_metric = 'e2e_latency_p95' if not df_ok['e2e_latency_p95'].dropna().empty else 'e2e_latency_mean'
    for rank, (_, row) in enumerate(df_ok.nsmallest(10, latency_metric).iterrows(), 1):
        rankings_data.append(['E2E Latency (Low)', rank, row['algorithm'], row[latency_metric]])

    if not df_ok['comm_score'].dropna().empty:
        for rank, (_, row) in enumerate(df_ok.nlargest(10, 'comm_score').iterrows(), 1):
            rankings_data.append(['Comm Score', rank, row['algorithm'], row['comm_score']])

    df_rankings = pd.DataFrame(rankings_data, columns=['Metric', 'Rank', 'Algorithm', 'Value'])
    for r_idx, row in enumerate(dataframe_to_rows(df_rankings, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    # Sheet 3: Summary Statistics
    ws3 = wb.create_sheet("Summary Stats")

    summary_data = []
    for metric in [
        'reward_mean',
        'e2e_latency_mean',
        'e2e_latency_p95',
        'deadline_miss_rate',
        'throughput_tasks_per_step',
        'comm_score',
        'latency_total',
        'energy_mean',
        'train_time_sec',
    ]:
        metric_data = df_ok[metric].dropna()
        if not metric_data.empty:
            summary_data.append([
                metric,
                metric_data.mean(),
                metric_data.std(),
                metric_data.min(),
                metric_data.max(),
                len(metric_data)
            ])

    df_summary = pd.DataFrame(summary_data, columns=['Metric', 'Mean', 'Std Dev', 'Min', 'Max', 'Count'])
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1 and c_idx > 1:
                cell.number_format = '0.0000'

    # Sheet 4: Algorithm Categories
    ws4 = wb.create_sheet("Categories")

    cat_data = []
    for category in ['On-Policy', 'Off-Policy']:
        df_cat = analyzer.df[(analyzer.df['category'] == category) & (analyzer.df['status'] == 'ok')]
        if not df_cat.empty:
            for algo in df_cat['algorithm']:
                cat_data.append([category, algo, 'Yes' if algo in ['MAPPO', 'QMIX', 'COMA', 'IPPO', 'VDN', 'MADDPG', 'IQL', 'MATD3'] else 'No'])

    df_cat = pd.DataFrame(cat_data, columns=['Category', 'Algorithm', 'Multi-Agent'])
    for r_idx, row in enumerate(dataframe_to_rows(df_cat, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws4.cell(row=r_idx, column=c_idx, value=value)

    # Sheet 5: Environment Comparison
    ws5 = wb.create_sheet("Environments")

    env_data = []
    for env in analyzer.df['environment'].unique():
        df_env = analyzer.df[(analyzer.df['environment'] == env) & (analyzer.df['status'] == 'ok')]
        if not df_env.empty:
            env_data.append([
                env,
                len(df_env),
                df_env['reward_mean'].mean(),
                df_env['reward_mean'].max(),
                df_env['e2e_latency_mean'].mean(),
                df_env['e2e_latency_mean'].min(),
                df_env['train_time_sec'].mean(),
            ])

    df_env_comp = pd.DataFrame(
        env_data,
        columns=['Environment', 'Count', 'Avg Reward', 'Best Reward', 'Avg E2E Latency', 'Best E2E Latency', 'Avg Time'],
    )
    for r_idx, row in enumerate(dataframe_to_rows(df_env_comp, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1 and c_idx > 2:
                cell.number_format = '0.0000'

    # Sheet 6: Efficiency Metrics
    ws6 = wb.create_sheet("Efficiency")

    df_efficiency = df_ok[
        ['algorithm', 'e2e_latency_mean', 'deadline_miss_rate', 'throughput_tasks_per_step', 'comm_score', 'reward_mean', 'train_time_sec', 'reward_per_sec']
    ].copy()
    sort_metric = 'comm_score' if not df_efficiency['comm_score'].dropna().empty else 'reward_per_sec'
    df_efficiency = df_efficiency.sort_values(sort_metric, ascending=False)

    for r_idx, row in enumerate(dataframe_to_rows(df_efficiency, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws6.cell(row=r_idx, column=c_idx, value=value)

    # Sheet 7: Diagnostic Report
    ws7 = wb.create_sheet("Diagnostics")

    ws7['A1'] = "Benchmark Diagnostics"
    ws7['A1'].font = Font(bold=True, size=12)

    row = 3
    ws7[f'A{row}'] = "Summary"
    ws7[f'A{row}'].font = Font(bold=True)

    row += 1
    summary = analyzer.get_success_summary()
    ws7[f'A{row}'] = f"Total Algorithms: {summary['total']}"
    row += 1
    ws7[f'A{row}'] = f"Successful: {summary['successful']}"
    row += 1
    ws7[f'A{row}'] = f"Failed: {summary['failed']}"
    row += 1
    ws7[f'A{row}'] = f"Success Rate: {summary['success_rate']}"

    row += 2
    ws7[f'A{row}'] = "Failed Runs"
    ws7[f'A{row}'].font = Font(bold=True)
    row += 1

    failed = analyzer.df[analyzer.df['status'] == 'failed']
    if not failed.empty:
        ws7[f'A{row}'] = "Algorithm"
        ws7[f'B{row}'] = "Environment"
        ws7[f'C{row}'] = "Error"
        row += 1

        for _, f_row in failed.iterrows():
            ws7[f'A{row}'] = f_row['algorithm']
            ws7[f'B{row}'] = f_row['environment']
            ws7[f'C{row}'] = f_row.get('error', 'Unknown')
            row += 1
    else:
        ws7[f'A{row}'] = "No failures"

    # Sheet 8: Composite Scores
    _add_composite_scores_sheet(analyzer, wb)

    # Save workbook
    wb.save(output_file)
    print(f"OK Excel analysis generated: {output_file}")


def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description="Generate RL-MEC benchmark reports")
    parser.add_argument("--input", type=Path, default=JSON_FILE, help="benchmark JSON input path")
    parser.add_argument("--output-md", type=Path, default=REPORT_FILE, help="markdown report output")
    parser.add_argument("--output-xlsx", type=Path, default=ANALYSIS_FILE, help="Excel analysis output")
    parser.add_argument("--figures-dir", type=Path, default=FIGURES_DIR, help="figures directory")
    parser.add_argument("--skip-excel", action="store_true", help="only generate markdown")
    args = parser.parse_args()

    print("=" * 60)
    print("RL-MEC Benchmark Report Generator")
    print("=" * 60)
    print()

    # Check input file exists
    if not args.input.exists():
        print(f"ERROR: {args.input} not found")
        return

    print(f"Input: {args.input}")
    print()

    # Create analyzer
    print("Processing benchmark data...")
    analyzer = BenchmarkAnalyzer(args.input)

    summary = analyzer.get_success_summary()
    print(f"[OK] Loaded {summary['total']} results ({summary['successful']} successful, {summary['failed']} failed)")
    print()

    # Generate markdown report
    print("Generating markdown report...")
    generate_markdown_report(analyzer, args.output_md, args.figures_dir)
    print()

    # Generate Excel analysis
    if not args.skip_excel:
        print("Generating Excel analysis...")
        generate_excel_analysis(analyzer, args.output_xlsx)
        print()

    # Summary
    print("=" * 60)
    print("[OK] Report generation complete!")
    print("=" * 60)
    print()
    print("Outputs:")
    print(f"  1. Markdown Report: {args.output_md}")
    if not args.skip_excel:
        print(f"  2. Excel Analysis: {args.output_xlsx}")
    print()

    # Show highlights
    print("Quick Highlights:")
    top_reward = analyzer.get_ranking('reward_mean', ascending=False, limit=3)
    if top_reward:
        print(f"  Top Reward: {top_reward[0]['algorithm']} ({top_reward[0]['value']:.4f})")

    top_time = analyzer.get_ranking('train_time_sec', ascending=True, limit=3)
    if top_time:
        print(f"  Fastest: {top_time[0]['algorithm']} ({top_time[0]['value']:.1f}s)")

    latency_rank_metric = 'e2e_latency_p95'
    if analyzer.df[latency_rank_metric].dropna().empty:
        latency_rank_metric = 'e2e_latency_mean'
    top_latency = analyzer.get_ranking(latency_rank_metric, ascending=True, limit=3)
    if top_latency:
        print(f"  Lowest E2E Latency: {top_latency[0]['algorithm']} ({top_latency[0]['value']:.4f})")

    top_comm = analyzer.get_ranking('comm_score', ascending=False, limit=3)
    if top_comm:
        print(f"  Best Comm Score: {top_comm[0]['algorithm']} ({top_comm[0]['value']:.4f})")


if __name__ == "__main__":
    main()
